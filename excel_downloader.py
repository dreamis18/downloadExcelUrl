import sys
import os
import requests
import openpyxl
from urllib.parse import urlparse
from pathvalidate import sanitize_filename

def safe_filename(name):
    """生成安全的文件名"""
    return sanitize_filename(str(name)).replace(' ', '_')[:200]

def get_extension(url):
    """从URL获取文件扩展名"""
    try:
        parsed = urlparse(url)
        path = parsed.path
        filename = os.path.basename(path)
        ext = os.path.splitext(filename)[1].split('?')[0][:10]
        return ext if ext else '.dat'
    except:
        return '.dat'

def download_file(url, save_path):
    """下载文件并保存到指定路径"""
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
        with requests.get(url, headers=headers, stream=True, timeout=30) as r:
            r.raise_for_status()
            with open(save_path, 'wb') as f:
                for chunk in r.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
        return True
    except Exception as e:
        print(f"下载失败: {str(e)}")
        return False

def process_excel(file_path):
    """处理Excel文件"""
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # 创建行文件夹
            folder_name = safe_filename(row[0].value)
            if not folder_name:
                continue
            
            os.makedirs(folder_name, exist_ok=True)

            # 处理每列数据
            for col_idx, cell in enumerate(row[1:], start=1):
                header = headers[col_idx]
                if not header:
                    continue
                
                # 分割多个URL
                urls = []
                if cell.value:
                    urls = [url.strip() for url in str(cell.value).split(',') if url.strip()]
                
                # 根据链接数量调整命名策略
                total_files = len(urls)
                for seq, url in enumerate(urls, start=1):
                    try:
                        safe_header = safe_filename(header)
                        ext = get_extension(url)
                        
                        # 单个文件不加序号，多个文件添加序号
                        if total_files > 1:
                            filename = f"{safe_header}{seq}{ext}"
                        else:
                            filename = f"{safe_header}{ext}"
                            
                        save_path = os.path.join(folder_name, filename)
                        
                        print(f"正在下载第{row_idx}行数据: {url}")
                        if download_file(url, save_path):
                            print(f"✓ 下载成功: {os.path.relpath(save_path)}")
                        else:
                            print(f"✗ 下载失败: {url}")
                    except Exception as e:
                        print(f"处理错误（行{row_idx}列{col_idx+1}）: {str(e)}")

    except Exception as e:
        print(f"Excel处理错误: {str(e)}")
    finally:
        input("按回车键退出...")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        input("请将Excel文件拖拽到本脚本上\n按回车键退出...")
    else:
        process_excel(sys.argv[1])